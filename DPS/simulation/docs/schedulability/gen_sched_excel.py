#!/usr/bin/env python3
"""Generate an Excel-friendly schedulability table from WCET terminal output.

What it does
- Parses lines like: "F1 physics iter  count=1140  max=631.500us  avg=..."
- Applies a configurable safety margin to convert max-observed time to analysis WCET $C$.
- Builds a modeled task set (leader + follower) consistent with docs/schedulability/task_model.md.
- Computes:
  - EDF utilization per task and total utilization
  - Fixed-priority response-time analysis (RTA) per task
- Writes output as:
  - .xlsx (if openpyxl is installed)
  - .csv (always)

Usage
  python3 docs/schedulability/gen_sched_excel.py --input wcet.txt

Notes
- The WCET you measured is thread CPU time. This is what we want for schedulability math.
- The follower model conservatively combines RX decode + FSM handling per message class.

Install optional xlsx support
  pip install openpyxl
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from math import ceil
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple


WCET_LINE_RE = re.compile(
    r"^(?P<name>.+?)\s+count=(?P<count>\d+)\s+max=(?P<max_us>[0-9.]+)us\s+avg=(?P<avg_us>[0-9.]+)us\s*$"
)


@dataclass(frozen=True)
class Observed:
    name: str
    count: int
    max_us: float
    avg_us: float


@dataclass(frozen=True)
class TaskRow:
    system: str  # leader|follower
    task: str
    components: str
    observed_us: float
    margin: float
    C_ms: float
    T_ms: float
    D_ms: float
    prio: int
    U: float
    R_ms: Optional[float]
    ok: Optional[bool]


def parse_wcet_text(text: str) -> Dict[str, Observed]:
    observed: Dict[str, Observed] = {}
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("==="):
            continue
        m = WCET_LINE_RE.match(line)
        if not m:
            continue
        name = m.group("name").strip()
        count = int(m.group("count"))
        max_us = float(m.group("max_us"))
        avg_us = float(m.group("avg_us"))
        observed[name] = Observed(name=name, count=count, max_us=max_us, avg_us=avg_us)
    return observed


def choose_margin(count: int, default_margin: float, rare_margin: float, rare_threshold: int) -> float:
    return rare_margin if count < rare_threshold else default_margin


@dataclass(frozen=True)
class Task:
    name: str
    C_ms: float
    T_ms: float
    D_ms: float
    prio: int  # smaller = higher priority


def utilization(tasks: Iterable[Task]) -> float:
    return sum(t.C_ms / t.T_ms for t in tasks)


def rta_fp(tasks: List[Task]) -> List[Tuple[Task, float, bool]]:
    tasks_sorted = sorted(tasks, key=lambda t: t.prio)
    out: List[Tuple[Task, float, bool]] = []

    for i, ti in enumerate(tasks_sorted):
        hp = tasks_sorted[:i]
        R = ti.C_ms
        for _ in range(100):
            interference = sum(ceil(R / tj.T_ms) * tj.C_ms for tj in hp)
            R_next = ti.C_ms + interference
            if abs(R_next - R) < 1e-9:
                R = R_next
                break
            R = R_next
            if R > ti.D_ms * 10:
                break
        ok = R <= ti.D_ms
        out.append((ti, R, ok))

    return out


def require(obs: Dict[str, Observed], key: str) -> Observed:
    if key not in obs:
        available = "\n".join(sorted(obs.keys()))
        raise SystemExit(
            f"Missing required WCET line: '{key}'.\n\nAvailable parsed names:\n{available}\n"
        )
    return obs[key]


def build_leader_rows(obs: Dict[str, Observed], default_margin: float, rare_margin: float, rare_threshold: int) -> Tuple[List[TaskRow], List[Task]]:
    # Map terminal names -> model tasks
    l1 = require(obs, "L1 tick-producer iter")
    l2 = require(obs, "L2 FSM per-event")
    l3 = require(obs, "L3 send one cmd")
    # L4/L5 may be absent in some runs; treat as optional
    l4 = obs.get("L4 rx msg->event")

    def c_ms(o: Observed) -> Tuple[float, float]:
        margin = choose_margin(o.count, default_margin, rare_margin, rare_threshold)
        return margin, (o.max_us / 1000.0) * margin

    rows: List[TaskRow] = []

    margin, C_tick = c_ms(l1)
    rows.append(
        TaskRow(
            system="leader",
            task="L_tick",
            components="L1 tick-producer iter",
            observed_us=l1.max_us,
            margin=margin,
            C_ms=C_tick,
            T_ms=250.0,
            D_ms=250.0,
            prio=6,
            U=C_tick / 250.0,
            R_ms=None,
            ok=None,
        )
    )

    margin, C_fsm = c_ms(l2)
    rows.append(
        TaskRow(
            system="leader",
            task="L_fsm",
            components="L2 FSM per-event",
            observed_us=l2.max_us,
            margin=margin,
            C_ms=C_fsm,
            T_ms=250.0,
            D_ms=250.0,
            prio=3,
            U=C_fsm / 250.0,
            R_ms=None,
            ok=None,
        )
    )

    margin, C_send = c_ms(l3)
    rows.append(
        TaskRow(
            system="leader",
            task="L_send",
            components="L3 send one cmd",
            observed_us=l3.max_us,
            margin=margin,
            C_ms=C_send,
            T_ms=250.0,
            D_ms=250.0,
            prio=4,
            U=C_send / 250.0,
            R_ms=None,
            ok=None,
        )
    )

    if l4 is not None:
        margin, C_rx = c_ms(l4)
        rows.append(
            TaskRow(
                system="leader",
                task="L_rx",
                components="L4 rx msg->event",
                observed_us=l4.max_us,
                margin=margin,
                C_ms=C_rx,
                T_ms=250.0,
                D_ms=250.0,
                prio=5,
                U=C_rx / 250.0,
                R_ms=None,
                ok=None,
            )
        )

    tasks = [Task(r.task, r.C_ms, r.T_ms, r.D_ms, r.prio) for r in rows]
    rta = {t.name: (R, ok) for t, R, ok in rta_fp(tasks)}
    rows2 = []
    for r in rows:
        R_ok = rta.get(r.task)
        rows2.append(
            TaskRow(
                **{**r.__dict__, "R_ms": (R_ok[0] if R_ok else None), "ok": (R_ok[1] if R_ok else None)}
            )
        )

    return rows2, tasks


def build_follower_rows(obs: Dict[str, Observed], default_margin: float, rare_margin: float, rare_threshold: int) -> Tuple[List[TaskRow], List[Task]]:
    f1 = require(obs, "F1 physics iter")
    f2 = require(obs, "F2 watchdog iter")
    f3 = require(obs, "F3 TCP msg->event")
    f4 = require(obs, "F4 UDP msg->event")

    # Event-type specific FSM paths may be missing if not triggered.
    ev_em = obs.get("F5 EVT_EMERGENCY")
    ev_to = obs.get("F5 EVT_LEADER_TIMEOUT")
    ev_cruise = obs.get("F5 EVT_CRUISE_CMD")
    ev_dist = obs.get("F5 EVT_DISTANCE")

    def c_ms(o: Observed) -> Tuple[float, float]:
        margin = choose_margin(o.count, default_margin, rare_margin, rare_threshold)
        return margin, (o.max_us / 1000.0) * margin

    # If per-event-type lines are absent, fall back to generic per-event.
    f5 = obs.get("F5 FSM per-event")

    def fallback_event_cost(label: str) -> Observed:
        if label in obs:
            return obs[label]
        if f5 is not None:
            return f5
        available = "\n".join(sorted(obs.keys()))
        raise SystemExit(
            f"Missing required FSM event cost for '{label}' and no 'F5 FSM per-event' found.\n\nAvailable:\n{available}\n"
        )

    ev_em = ev_em or fallback_event_cost("F5 EVT_EMERGENCY")
    ev_to = ev_to or fallback_event_cost("F5 EVT_LEADER_TIMEOUT")
    ev_cruise = ev_cruise or fallback_event_cost("F5 EVT_CRUISE_CMD")
    ev_dist = ev_dist or fallback_event_cost("F5 EVT_DISTANCE")

    # Conservative combination: decode+enqueue + FSM handling
    margin_f4, C_f4 = c_ms(f4)
    margin_f3, C_f3 = c_ms(f3)

    margin_em, C_ev_em = c_ms(ev_em)
    margin_to, C_ev_to = c_ms(ev_to)
    margin_cruise, C_ev_cruise = c_ms(ev_cruise)
    margin_dist, C_ev_dist = c_ms(ev_dist)

    # For combined tasks, use the max margin among components (conservative).
    def combine(components: List[Tuple[Observed, float, float]]) -> Tuple[str, float, float]:
        names = "+".join(o.name for o, _, _ in components)
        margin = max(m for _, m, _ in components)
        C_ms_sum = sum(C for _, _, C in components)
        observed_sum_us = sum(o.max_us for o, _, _ in components)
        return names, margin, C_ms_sum, observed_sum_us

    rows: List[TaskRow] = []

    # Emergency: F4 + EVT_EMERGENCY
    comps = [(f4, margin_f4, C_f4), (ev_em, margin_em, C_ev_em)]
    comp_names, margin, C_em, obs_us = combine(comps)
    rows.append(
        TaskRow(
            system="follower",
            task="F_emergency",
            components=comp_names,
            observed_us=obs_us,
            margin=margin,
            C_ms=C_em,
            T_ms=50.0,
            D_ms=50.0,
            prio=1,
            U=C_em / 50.0,
            R_ms=None,
            ok=None,
        )
    )

    # Watchdog enqueue loop
    margin, C_watchdog = c_ms(f2)
    rows.append(
        TaskRow(
            system="follower",
            task="F_watchdog",
            components=f2.name,
            observed_us=f2.max_us,
            margin=margin,
            C_ms=C_watchdog,
            T_ms=100.0,
            D_ms=100.0,
            prio=2,
            U=C_watchdog / 100.0,
            R_ms=None,
            ok=None,
        )
    )

    # Timeout handling in FSM (treat as periodic to match watchdog period; pessimistic)
    margin, C_timeout = c_ms(ev_to)
    rows.append(
        TaskRow(
            system="follower",
            task="F_timeout_fsm",
            components=ev_to.name,
            observed_us=ev_to.max_us,
            margin=margin,
            C_ms=C_timeout,
            T_ms=100.0,
            D_ms=100.0,
            prio=2,
            U=C_timeout / 100.0,
            R_ms=None,
            ok=None,
        )
    )

    # Cruise: F3 + EVT_CRUISE_CMD
    comps = [(f3, margin_f3, C_f3), (ev_cruise, margin_cruise, C_ev_cruise)]
    comp_names, margin, C_cruise, obs_us = combine(comps)
    rows.append(
        TaskRow(
            system="follower",
            task="F_cruise",
            components=comp_names,
            observed_us=obs_us,
            margin=margin,
            C_ms=C_cruise,
            T_ms=250.0,
            D_ms=250.0,
            prio=5,
            U=C_cruise / 250.0,
            R_ms=None,
            ok=None,
        )
    )

    # Distance: F4 + EVT_DISTANCE
    comps = [(f4, margin_f4, C_f4), (ev_dist, margin_dist, C_ev_dist)]
    comp_names, margin, C_dist, obs_us = combine(comps)
    rows.append(
        TaskRow(
            system="follower",
            task="F_distance",
            components=comp_names,
            observed_us=obs_us,
            margin=margin,
            C_ms=C_dist,
            T_ms=250.0,
            D_ms=250.0,
            prio=5,
            U=C_dist / 250.0,
            R_ms=None,
            ok=None,
        )
    )

    # Physics loop
    margin, C_phys = c_ms(f1)
    rows.append(
        TaskRow(
            system="follower",
            task="F_physics",
            components=f1.name,
            observed_us=f1.max_us,
            margin=margin,
            C_ms=C_phys,
            T_ms=250.0,
            D_ms=250.0,
            prio=8,
            U=C_phys / 250.0,
            R_ms=None,
            ok=None,
        )
    )

    tasks = [Task(r.task, r.C_ms, r.T_ms, r.D_ms, r.prio) for r in rows]
    rta = {t.name: (R, ok) for t, R, ok in rta_fp(tasks)}
    rows2 = []
    for r in rows:
        R_ok = rta.get(r.task)
        rows2.append(
            TaskRow(
                **{**r.__dict__, "R_ms": (R_ok[0] if R_ok else None), "ok": (R_ok[1] if R_ok else None)}
            )
        )

    return rows2, tasks


def write_csv(path: Path, rows: List[TaskRow]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "system",
                "task",
                "components",
                "observed_us",
                "margin",
                "C_ms",
                "T_ms",
                "D_ms",
                "prio",
                "U",
                "R_ms",
                "OK",
            ]
        )
        for r in rows:
            w.writerow(
                [
                    r.system,
                    r.task,
                    r.components,
                    f"{r.observed_us:.3f}",
                    f"{r.margin:.3f}",
                    f"{r.C_ms:.6f}",
                    f"{r.T_ms:.3f}",
                    f"{r.D_ms:.3f}",
                    r.prio,
                    f"{r.U:.8f}",
                    "" if r.R_ms is None else f"{r.R_ms:.6f}",
                    "" if r.ok is None else ("YES" if r.ok else "NO"),
                ]
            )


def write_xlsx(path: Path, rows: List[TaskRow], leader_tasks: List[Task], follower_tasks: List[Task]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except Exception:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"

    headers = [
        "system",
        "task",
        "components",
        "observed_us",
        "margin",
        "C_ms",
        "T_ms",
        "D_ms",
        "prio",
        "U (=C/T)",
        "R_ms (RTA)",
        "OK (R<=D)",
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in rows:
        ws.append(
            [
                r.system,
                r.task,
                r.components,
                r.observed_us,
                r.margin,
                r.C_ms,
                r.T_ms,
                r.D_ms,
                r.prio,
                r.U,
                r.R_ms if r.R_ms is not None else "",
                "YES" if r.ok else ("NO" if r.ok is False else ""),
            ]
        )

    # Simple summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["System", "Total U", "EDF (implicit) pass if U<=1"])
    ws2.append(["Leader", utilization(leader_tasks), "PASS" if utilization(leader_tasks) <= 1.0 else "FAIL"])
    ws2.append(["Follower", utilization(follower_tasks), "PASS" if utilization(follower_tasks) <= 1.0 else "FAIL"])

    for c in range(1, 4):
        ws2.cell(row=1, column=c).font = header_font

    # Column widths
    widths = {
        1: 10,
        2: 18,
        3: 44,
        4: 12,
        5: 10,
        6: 12,
        7: 10,
        8: 10,
        9: 6,
        10: 12,
        11: 12,
        12: 10,
    }
    for col, w in widths.items():
        ws.column_dimensions[chr(ord('A') + col - 1)].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return True


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Path to text file containing pasted WCET output")
    ap.add_argument("--out", default="docs/schedulability/schedulability", help="Output base path (without extension)")
    ap.add_argument("--default-margin", type=float, default=2.0, help="Margin for well-sampled counters")
    ap.add_argument("--rare-margin", type=float, default=3.0, help="Margin for rare counters")
    ap.add_argument("--rare-threshold", type=int, default=100, help="If count<threshold, treated as rare")
    args = ap.parse_args()

    text = Path(args.input).read_text(encoding="utf-8", errors="replace")
    obs = parse_wcet_text(text)

    leader_rows, leader_tasks = build_leader_rows(
        obs, args.default_margin, args.rare_margin, args.rare_threshold
    )
    follower_rows, follower_tasks = build_follower_rows(
        obs, args.default_margin, args.rare_margin, args.rare_threshold
    )

    all_rows = leader_rows + follower_rows

    base = Path(args.out)
    csv_path = base.with_suffix(".csv")
    xlsx_path = base.with_suffix(".xlsx")

    write_csv(csv_path, all_rows)

    xlsx_ok = write_xlsx(xlsx_path, all_rows, leader_tasks, follower_tasks)

    print(f"Wrote: {csv_path}")
    if xlsx_ok:
        print(f"Wrote: {xlsx_path}")
    else:
        print("openpyxl not installed; skipped .xlsx (CSV is still usable in Excel).")
        print("To enable .xlsx: pip install openpyxl")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
